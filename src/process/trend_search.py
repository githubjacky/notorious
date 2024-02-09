import datetime
from functools import cached_property
from loguru import logger
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from typing import List, Literal, Generator, Optional

from .utils import get_target_list
from gtab import BaseTrendSearch


class TrendSearch(BaseTrendSearch):
    def __init__(self,
                 geo: str = "",
                 period: List[str] = ["2016-01-01", "2018-07-31"]):
        super().__init__(geo, period)


    @cached_property
    def __begin_date(self) -> datetime.date:
        return datetime.datetime.strptime(
            self.period[0], '%Y-%m-%d'
        ).date()


    @cached_property
    def __end_date(self) -> datetime.date:
        return datetime.datetime.strptime(
            self.period[1], '%Y-%m-%d'
        ).date()


    @property
    def __begin_year(self) -> int:
        return self.__begin_date.year


    @property
    def __begin_month(self) -> int:
        return self.__begin_date.month


    @property
    def __end_year(self):
        return self.__end_date.year


    @property
    def __end_month(self):
        return self.__end_date.month


    @staticmethod
    def __add_zero(x: int) -> str:
        return f'0{x}' if len(str(x))==1 else str(x)


    def __inner_col_generator_by_sum(self,
                                     curr_year: int,
                                     begin_month: int = 1,
                                     end_month: int = 12
                                    ) -> Generator[tuple, None, None]:
            curr_month = begin_month
            while curr_month <= end_month:
                col = []
                for name in self.names:
                    item = {
                        'year': curr_year,
                        'month': curr_month,
                        'name': name,
                        'suffix': self.suffix
                    }

                    if self.collection.count_documents(item) != 0:
                        col.append(np.sum(
                            [
                                i['max_ratio']
                                for i in self.collection.find(item)
                            ]
                        ))
                    else:
                        col.append(None)

                yield f'{curr_year}_{self.__add_zero(curr_month)}', col
                curr_month += 1


    def __col_generator_by_sum(self) -> Generator[tuple, None, None]:
        """Generate the columns of the excel sheet, transforming the weekly
        frequency data to month through summation.
        """
        yield from self.__inner_col_generator_by_sum(self.__begin_year, self.__begin_month)

        curr_year = self.__begin_year + 1
        while curr_year < self.__end_year:
            yield from self.__inner_col_generator_by_sum(curr_year)
            curr_year += 1

        yield from self.__inner_col_generator_by_sum(curr_year, 1, self.__end_month)


    def __inner_col_generator_by_mean(self,
                                     curr_year: int,
                                     begin_month: int = 1,
                                     end_month: int = 12
                                    ) -> Generator[tuple, None, None]:
            curr_month = begin_month
            while curr_month <= end_month:
                col = []
                for name in self.names:
                    item = {
                        'year': curr_year,
                        'month': curr_month,
                        'name': name,
                        'suffix': self.suffix
                    }

                    if self.collection.count_documents(item) != 0:
                        col.append(np.mean(
                            [
                                i['max_ratio']
                                for i in self.collection.find(item)
                            ]
                        ))
                    else:
                        col.append(None)

                yield f'{curr_year}_{self.__add_zero(curr_month)}', col
                curr_month += 1


    def __col_generator_by_mean(self) -> Generator[tuple, None, None]:
        """Generate the columns of the excel sheet, transforming the weekly
        frequency data to month through taking average.
        """
        yield from self.__inner_col_generator_by_mean(self.__begin_year, self.__begin_month)

        curr_year = self.__begin_year + 1
        while curr_year < self.__end_year:
            yield from self.__inner_col_generator_by_mean(curr_year)
            curr_year += 1

        yield from self.__inner_col_generator_by_mean(curr_year, 1, self.__end_month)


    def __merge_res(self,
                    df_name: pd.DataFrame,
                    df: pd.DataFrame,
                    raw_data_path: str,
                    sheet_name: str,
                    min_value: float = 0.00001) -> pd.DataFrame:
        """Retain the old sheet and append non-duplicated columns.
        """

        wb = load_workbook(raw_data_path, read_only=True)
        if sheet_name in wb.sheetnames:
            base_df = (
                pd
                .read_excel(raw_data_path, sheet_name=sheet_name)
                .drop(columns=df_name.columns)
            )
            for col in df.columns:
                if base_df.get(col) is None:
                    base_df[col] = df[col]
            base_df = base_df.reindex(sorted(base_df.columns), axis=1)

            output_df = pd.concat([df_name, base_df], axis=1)
        else:
            output_df = pd.concat([df_name, df], axis=1)

        output_df.replace(0, min_value, inplace=True)
        with pd.ExcelWriter(path = raw_data_path,
                            mode = 'a',
                            if_sheet_exists = 'replace'
                            ) as writer:

            output_df.to_excel(
                writer,
                sheet_name = sheet_name,
                index = False
            )
        return output_df


    @staticmethod
    def create_ri_for_regression(predator_victim_sheet_name: str = '名單',
                                 ri_sheet_name: str = 'new_Ri_sum_smooth_addmin',
                                 new_sheet_name: str = 'new_Ri_for_regression',
                                 raw_data_path: str = 'data/raw/victim_list_11222023.xlsx'
                                 ) -> pd.DataFrame:
        df_predator_victim = pd.read_excel(raw_data_path, predator_victim_sheet_name)
        df_ri = pd.read_excel(raw_data_path, sheet_name=ri_sheet_name)
        predator2id = {
            df_ri.iloc[i]['predator']: df_ri.iloc[i]['predator_id']
            for i in range(len(df_ri))
        }
        df = pd.DataFrame({
            'predator': df_predator_victim['predator'],
            'predator_id': [predator2id[df_predator_victim.iloc[i]['predator']] for i in range(len(df_predator_victim))],
            'victim': df_predator_victim['victim'],
            'victim_id': range(1, len(df_predator_victim)+1)
        })

        months = df_ri.columns.drop(['predator', 'predator_id'])
        for month in months:
            predaotr2ri = {
                df_ri.iloc[i]['predator']: df_ri.iloc[i][month]
                for i in range(len(df_ri))
            }
            df[month] = [
                predaotr2ri[predator]
                for predator in df['predator']
            ]
        with pd.ExcelWriter(raw_data_path,
                            mode='a',
                            if_sheet_exists = 'replace'
                            ) as writer:
            df.to_excel(writer, sheet_name = new_sheet_name, index = False)
        return df


    def calibrate(self,
                  db_name: str,
                  collection_name: str,
                  continuous_mode: Literal[True, False] = True,
                  max_retry: int = 5,
                  target: Literal['predator', 'victim'] = 'predator',
                  merge_method: Literal['sum', 'mean'] = 'sum',
                  raw_data_path: str = 'data/raw/victim_list_11222023.xlsx',
                  sheet_name: str = 'new_Ri_sum_smooth',
                  min_value: float = 0.00001,
                  predator_victim_sheet_name: Optional[str] = None,
                  new_Ri_for_regression_sheet_name: Optional[str] = None
                  ) -> pd.DataFrame:
        """Calibrate and merge the results, appending/creating a excel sheet.

        Args:
            `db_name`:          name of the database
            `collection_name`:  name of the collection
            `continuous_mode`:  whether to keep calibrate bad keywods which
                fail to calibrate.
            `max_retry`:        maximum round of tries to calibrate bad keywords
            `target`:           whether to calibrate the 'victim' or 'predator'
            `merge_method`:     transforming the data through 'summation' or
                'taking average'
            `raw_data_path`:    the path of the excel
            `sheet_name`:       sheet name for recording calibration results
            `min_value`:        if fail to calibration results, or the trends 
                is smaller than `min_value`), replace them with `min_value`

        Returns:
            None
        """
        self.names = get_target_list(target)
        self.calibrate_batch(
            self.names,
            db_name,
            collection_name,
            continuous_mode,
            max_retry
        )
        # self.collection = self.client[db_name][collection_name]

        df = pd.DataFrame(
            {
                month: col
                for month, col in self.__col_generator_by_sum()
            }
            if merge_method == 'sum'
            else
            {
                month: col
                for month, col in self.__col_generator_by_mean()
            }
        )
        df_name = pd.DataFrame({
            target: self.names,
            f'{target}_id': [i+1 for i in range(len(self.names))]
        })
        output_df = self.__merge_res(df_name, df, raw_data_path, sheet_name, min_value)

        if target == 'predator':
            logger.info('create ri for regression')
            self.create_ri_for_regression(
                predator_victim_sheet_name if predator_victim_sheet_name is not None else '名單',
                sheet_name,
                new_Ri_for_regression_sheet_name if new_Ri_for_regression_sheet_name is not None else 'new_Ri_for_regression_addmin',
                raw_data_path
            )

        return output_df
